from dataclasses import asdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


EXPORTS_DIR = Path("exports")


def exportar_excel(leads):
    # Gera uma planilha simples para análise e uso fora do programa.
    EXPORTS_DIR.mkdir(exist_ok=True)
    caminho = EXPORTS_DIR / "leads.xlsx"
    colunas = [
        "nome",
        "telefone",
        "cidade",
        "nicho",
        "endereco",
        "site",
        "instagram",
        "filial",
        "motivo_filial",
        "nota_google",
        "quantidade_avaliacoes",
        "score_contato",
        "score_presenca_digital",
        "score_oportunidade",
        "score_qualidade_dados",
        "score",
        "classificacao",
        "motivos_score",
        "mensagem_gerada",
        "status",
    ]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Leads"
    headers = [
        "Nome",
        "Telefone",
        "Cidade",
        "Nicho",
        "Endereço",
        "Site",
        "Instagram",
        "Filial",
        "Motivo Filial",
        "Nota Google",
        "Dados OSM",
        "Score Contato",
        "Score Presença Digital",
        "Score Oportunidade",
        "Score Qualidade Dados",
        "Score Final",
        "Classificação",
        "Motivos Score",
        "Mensagem Gerada",
        "Status",
    ]
    sheet.append(headers)

    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="111111")

    for lead in leads:
        dados = asdict(lead)
        dados["motivos_score"] = " | ".join(lead.motivos_score)
        sheet.append([dados[coluna] for coluna in colunas])

    for column in sheet.columns:
        column_letter = column[0].column_letter
        max_length = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[column_letter].width = min(max_length + 2, 56)

    workbook.save(caminho)
    return caminho
