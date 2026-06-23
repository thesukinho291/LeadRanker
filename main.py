from dataclasses import asdict, dataclass, field
from pathlib import Path
from urllib.parse import quote
import re
import threading
import time
import webbrowser

try:
    import customtkinter as ctk
    import requests
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from PIL import Image
except ModuleNotFoundError as error:
    print(f"Pacote não encontrado: {error.name}")
    print("Instale as dependências com: python -m pip install -r requirements.txt")
    raise SystemExit(1)


APP_TITLE = "LeadRanker"
EXPORTS_DIR = Path("exports")
IMAGES_DIR = Path("imagens")
LOGO_TEXT_PATH = IMAGES_DIR / "Logo com Texto.png"
NOMINATIM_URL = "https://nominatim.openstreetmap.org/search"
OVERPASS_URL = "https://overpass-api.de/api/interpreter"
OSM_USER_AGENT = "LeadRankerAI/1.0 (portfolio project)"

COLORS = {
    "bg": "#060606",
    "panel": "#0b0b0b",
    "panel_alt": "#111111",
    "panel_hover": "#1a1a1a",
    "border": "#2e2e2e",
    "text": "#f5f5f5",
    "muted": "#8f8f8f",
    "muted_2": "#b8b8b8",
    "accent": "#ffffff",
    "accent_text": "#050505",
    "success": "#47d18c",
}

CLASS_COLORS = {
    "Quente": "#47d18c",
    "Bom": "#b7e64a",
    "Médio": "#f2c94c",
    "Frio": "#9ca3af",
}

STATUSES = ["Novo", "Chamado", "Respondeu", "Fechado", "Sem interesse"]
FILTER_OPTIONS = ["Ambos", "Sim", "Não"]
CHAIN_BRANDS = [
    "drogasil",
    "droga raia",
    "raia drogasil",
    "pague menos",
    "drogaria sao paulo",
    "drogaria são paulo",
    "drogaria araujo",
    "drogaria araújo",
    "carrefour",
    "extra",
    "assai",
    "assaí",
    "atacadao",
    "atacadão",
    "mcdonald",
    "burger king",
    "subway",
    "habib",
    "cacaushow",
    "cacau show",
    "o boticario",
    "o boticário",
    "renner",
    "riachuelo",
    "magazine luiza",
]


@dataclass
class Lead:
    nome: str
    telefone: str
    cidade: str
    nicho: str
    endereco: str
    site: str
    nota_google: float
    quantidade_avaliacoes: int
    site_basico: bool
    instagram: str = ""
    filial: bool = False
    motivo_filial: str = ""
    status: str = "Novo"
    score: int = 0
    score_contato: int = 0
    score_presenca_digital: int = 0
    score_oportunidade: int = 0
    score_qualidade_dados: int = 0
    classificacao: str = "Frio"
    mensagem_gerada: str = ""
    motivos_score: list[str] = field(default_factory=list)


def classificar_lead(score):
    if score >= 70:
        return "Quente"
    if score >= 50:
        return "Bom"
    if score >= 30:
        return "Médio"
    return "Frio"


def calcular_score_analitico(lead):
    # Divide o score em partes para deixar a análise mais fácil de explicar.
    score_contato, motivos_contato = calcular_score_contato(lead)
    score_presenca, motivos_presenca = calcular_score_presenca_digital(lead)
    score_oportunidade, motivos_oportunidade = calcular_score_oportunidade(lead)
    score_dados, motivos_dados = calcular_score_qualidade_dados(lead)

    score_final = round(
        (score_contato * 0.25)
        + (score_presenca * 0.35)
        + (score_oportunidade * 0.30)
        + (score_dados * 0.10)
    )

    motivos = motivos_contato + motivos_presenca + motivos_oportunidade + motivos_dados

    if lead.filial:
        score_final = max(score_final - 35, 0)
        motivos.append(f"Possível filial/rede: {lead.motivo_filial}")

    lead.score_contato = score_contato
    lead.score_presenca_digital = score_presenca
    lead.score_oportunidade = score_oportunidade
    lead.score_qualidade_dados = score_dados
    lead.motivos_score = motivos[:8]

    return min(score_final, 100)


def calcular_score_contato(lead):
    score = 0
    motivos = []

    if lead.telefone:
        score += 70
        motivos.append("Possui telefone para contato.")
    else:
        motivos.append("Não possui telefone cadastrado.")

    if lead.endereco and lead.endereco != "Endereço não informado":
        score += 20
        motivos.append("Possui endereço cadastrado.")

    if lead.site or lead.instagram:
        score += 10

    return min(score, 100), motivos


def calcular_score_presenca_digital(lead):
    score = 0
    motivos = []

    if not lead.site:
        score += 45
        motivos.append("Não possui site cadastrado.")
    elif not lead.site.startswith("https://"):
        score += 25
        motivos.append("Site cadastrado não usa HTTPS.")

    if not lead.instagram:
        score += 30
        motivos.append("Não possui Instagram cadastrado no OpenStreetMap.")

    if lead.site_basico:
        score += 25
        motivos.append("Site aparenta ser básico, lento ou instável.")

    return min(score, 100), motivos


def calcular_score_oportunidade(lead):
    score = 0
    motivos = []

    if not lead.filial:
        score += 30
        motivos.append("Não parece ser uma grande rede.")

    if lead.telefone:
        score += 20

    if not lead.site:
        score += 25

    if not lead.instagram:
        score += 15

    if lead.site_basico:
        score += 10

    return min(score, 100), motivos


def calcular_score_qualidade_dados(lead):
    score = min(lead.quantidade_avaliacoes * 12, 70)
    motivos = [f"Cadastro possui {lead.quantidade_avaliacoes} dado(s) útil(eis) no OSM."]

    if lead.endereco and lead.endereco != "Endereço não informado":
        score += 15

    if lead.telefone:
        score += 15

    return min(score, 100), motivos


def gerar_mensagem(lead):
    oportunidade = (
        "melhorar a presença digital"
        if lead.site
        else "criar uma presença digital mais forte para captar clientes locais"
    )

    return (
        "Boa tarde, tudo bem? Estava pesquisando sobre a empresa de vocês "
        f"e percebi uma oportunidade de {oportunidade}. Trabalho com soluções "
        "simples para empresas e posso te mostrar uma ideia rápida, sem compromisso."
    )


def processar_lead(lead):
    lead.score = calcular_score_analitico(lead)
    lead.classificacao = classificar_lead(lead.score)
    lead.mensagem_gerada = gerar_mensagem(lead)
    return lead


def buscar_leads_openstreetmap(cidade, nicho, quantidade):
    # Busca a cidade primeiro para limitar a consulta ao mapa daquela região.
    quantidade = max(1, min(quantidade, 50))
    bbox = buscar_bbox_cidade(cidade)
    consulta = montar_consulta_overpass(nicho, bbox, quantidade)

    resposta = requests.post(
        OVERPASS_URL,
        data={"data": consulta},
        timeout=45,
        headers={"User-Agent": OSM_USER_AGENT},
    )
    resposta.raise_for_status()
    dados = resposta.json()

    leads = []
    vistos = set()

    for elemento in dados.get("elements", []):
        if len(leads) >= quantidade:
            break

        tags = elemento.get("tags", {})
        nome = tags.get("name", "").strip()
        if not nome:
            continue

        chave = normalizar_texto(
            f"{nome}-{tags.get('addr:street', '')}-{tags.get('phone', '')}"
        )
        if chave in vistos:
            continue

        vistos.add(chave)
        leads.append(processar_lead(criar_lead_osm(elemento, cidade, nicho)))

    return sorted(leads, key=lambda item: item.score, reverse=True)


def buscar_bbox_cidade(cidade):
    resposta = requests.get(
        NOMINATIM_URL,
        params={
            "q": f"{cidade}, Brasil",
            "format": "json",
            "limit": 1,
            "addressdetails": 1,
        },
        timeout=20,
        headers={"User-Agent": OSM_USER_AGENT},
    )
    resposta.raise_for_status()
    resultados = resposta.json()

    if not resultados:
        raise RuntimeError(f"Não encontrei a cidade '{cidade}' no OpenStreetMap.")

    sul, norte, oeste, leste = resultados[0]["boundingbox"]
    return float(sul), float(oeste), float(norte), float(leste)


def montar_consulta_overpass(nicho, bbox, quantidade):
    # Monta a consulta que o Overpass usa para encontrar empresas no OpenStreetMap.
    sul, oeste, norte, leste = bbox
    blocos = []

    for filtro in filtros_osm_por_nicho(nicho):
        blocos.append(f"node{filtro}({sul},{oeste},{norte},{leste});")
        blocos.append(f"way{filtro}({sul},{oeste},{norte},{leste});")
        blocos.append(f"relation{filtro}({sul},{oeste},{norte},{leste});")

    limite = max(quantidade * 5, 40)
    return f"""
[out:json][timeout:35];
(
  {' '.join(blocos)}
);
out tags center {limite};
"""


def filtros_osm_por_nicho(nicho):
    texto = normalizar_texto(nicho)
    filtros = []

    mapas = [
        (["restaurante", "lanchonete", "pizza", "hamburguer", "cafe"], '["amenity"~"restaurant|fast_food|cafe|bar|pub"]'),
        (["clinica", "medico", "dentista", "saude"], '["healthcare"~"clinic|doctor|dentist|physiotherapist"]'),
        (["clinica", "medico", "dentista", "saude"], '["amenity"~"clinic|doctors|dentist"]'),
        (["estetica", "beleza", "salao", "barbearia"], '["shop"~"beauty|hairdresser|cosmetics"]'),
        (["academia", "fitness", "musculacao"], '["leisure"="fitness_centre"]'),
        (["pet", "veterinario", "veterinaria"], '["shop"="pet"]'),
        (["pet", "veterinario", "veterinaria"], '["amenity"="veterinary"]'),
        (["mercado", "supermercado", "padaria", "acougue"], '["shop"~"supermarket|convenience|bakery|butcher"]'),
        (["farmacia", "drogaria"], '["amenity"="pharmacy"]'),
        (["farmacia", "drogaria"], '["shop"="chemist"]'),
        (["hotel", "pousada"], '["tourism"~"hotel|guest_house|hostel"]'),
        (["escola", "curso"], '["amenity"~"school|college|language_school"]'),
        (["loja", "comercio"], '["shop"]'),
    ]

    for palavras, filtro in mapas:
        if any(palavra in texto for palavra in palavras):
            filtros.append(filtro)

    if not filtros:
        termo = re.escape(nicho.strip())
        filtros.append(f'["name"~"{termo}",i]')

    return list(dict.fromkeys(filtros))


def criar_lead_osm(elemento, cidade, nicho):
    tags = elemento.get("tags", {})
    site = tags.get("website") or tags.get("contact:website") or ""
    telefone = tags.get("phone") or tags.get("contact:phone") or tags.get("mobile") or ""
    filial, motivo_filial = detectar_filial(tags)

    return Lead(
        nome=tags.get("name", "Empresa sem nome"),
        telefone=telefone,
        cidade=cidade.title(),
        nicho=nicho.title(),
        endereco=montar_endereco_osm(tags),
        site=site,
        nota_google=0,
        quantidade_avaliacoes=contar_dados_osm(tags),
        site_basico=bool(site and not site.startswith("https://")),
        instagram=extrair_instagram(tags),
        filial=filial,
        motivo_filial=motivo_filial,
    )


def extrair_instagram(tags):
    instagram = (
        tags.get("contact:instagram")
        or tags.get("instagram")
        or tags.get("social_media:instagram")
        or ""
    )
    return instagram.strip()


def detectar_filial(tags):
    # Tenta identificar redes grandes para evitar priorizar empresas que já têm estrutura.
    nome = normalizar_texto(tags.get("name", ""))
    marca = normalizar_texto(tags.get("brand", ""))
    operador = normalizar_texto(tags.get("operator", ""))

    for campo, valor in [("nome", nome), ("marca", marca), ("operador", operador)]:
        for rede in CHAIN_BRANDS:
            if normalizar_texto(rede) in valor:
                return True, f"{campo} contém '{rede}'"

    if tags.get("brand:wikidata") or tags.get("brand:wikipedia"):
        return True, "possui marca cadastrada no OSM"

    if normalizar_texto(tags.get("branch", "")) in ["yes", "sim"]:
        return True, "tag branch indica filial"

    return False, ""


def montar_endereco_osm(tags):
    rua = tags.get("addr:street", "")
    numero = tags.get("addr:housenumber", "")
    bairro = tags.get("addr:suburb") or tags.get("addr:neighbourhood") or ""
    cidade = tags.get("addr:city", "")

    partes = []
    if rua:
        partes.append(f"{rua}, {numero}".strip().strip(","))
    if bairro:
        partes.append(bairro)
    if cidade:
        partes.append(cidade)

    return " - ".join(partes) or "Endereço não informado"


def contar_dados_osm(tags):
    campos = [
        "name",
        "phone",
        "contact:phone",
        "website",
        "contact:website",
        "addr:street",
        "addr:housenumber",
        "opening_hours",
    ]
    return sum(1 for campo in campos if tags.get(campo))


def normalizar_texto(texto):
    substituicoes = str.maketrans(
        "áàãâéêíóôõúçÁÀÃÂÉÊÍÓÔÕÚÇ",
        "aaaaeeioooucAAAAEEIOOOUC",
    )
    return texto.translate(substituicoes).lower().strip()


def analisar_sites(leads):
    for lead in leads:
        if lead.site:
            lead.site_basico = analisar_site_basico(lead.site)
        processar_lead(lead)
    return sorted(leads, key=lambda item: item.score, reverse=True)


def analisar_site_basico(site):
    # Faz uma checagem simples para descobrir se o site parece antigo ou instável.
    if not site:
        return False

    if not site.startswith("https://"):
        return True

    try:
        inicio = time.perf_counter()
        resposta = requests.get(site, timeout=8, headers={"User-Agent": OSM_USER_AGENT})
        tempo = time.perf_counter() - inicio
    except requests.RequestException:
        return True

    return resposta.status_code >= 400 or tempo > 3


def exportar_excel(leads):
    # Gera uma planilha simples para análise e uso fora do programa.
    EXPORTS_DIR.mkdir(exist_ok=True)
    caminho = EXPORTS_DIR / "leads.xlsx"
    colunas = [
        "nome",
        "telefone",
        "cidade",
        "nicho",
        "endereco",
        "site",
        "instagram",
        "filial",
        "motivo_filial",
        "nota_google",
        "quantidade_avaliacoes",
        "score_contato",
        "score_presenca_digital",
        "score_oportunidade",
        "score_qualidade_dados",
        "score",
        "classificacao",
        "motivos_score",
        "mensagem_gerada",
        "status",
    ]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Leads"
    headers = [coluna.replace("_", " ").title() for coluna in colunas]
    headers = [
        "Nome",
        "Telefone",
        "Cidade",
        "Nicho",
        "Endereço",
        "Site",
        "Instagram",
        "Filial",
        "Motivo Filial",
        "Nota Google",
        "Dados OSM",
        "Score Contato",
        "Score Presença Digital",
        "Score Oportunidade",
        "Score Qualidade Dados",
        "Score Final",
        "Classificação",
        "Motivos Score",
        "Mensagem Gerada",
        "Status",
    ]
    sheet.append(headers)

    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="111111")

    for lead in leads:
        dados = asdict(lead)
        dados["motivos_score"] = " | ".join(lead.motivos_score)
        sheet.append([dados[coluna] for coluna in colunas])

    for column in sheet.columns:
        column_letter = column[0].column_letter
        max_length = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[column_letter].width = min(max_length + 2, 56)

    workbook.save(caminho)
    return caminho


class LeadRankerApp(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.leads = []
        self.selected_index = None
        self.status_filter = "Todos"
        self.site_filter = "Ambos"
        self.instagram_filter = "Ambos"
        self.telefone_filter = "Ambos"
        self.is_loading = False
        self.logo_text_image = None

        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("dark-blue")

        self.title(APP_TITLE)
        self.geometry("1280x760")
        self.minsize(1120, 700)
        self.configure(fg_color=COLORS["bg"])
        self.attributes("-alpha", 0.96)
        self._load_visual_assets()

        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)

        self.sidebar = ctk.CTkFrame(self, width=300, fg_color=COLORS["bg"], corner_radius=0)
        self.sidebar.grid(row=0, column=0, sticky="nsew")
        self.sidebar.grid_propagate(False)
        self.sidebar.grid_columnconfigure(0, weight=1)
        self.sidebar.grid_rowconfigure(1, weight=1)

        self.main_area = ctk.CTkFrame(self, fg_color=COLORS["bg"], corner_radius=0)
        self.main_area.grid(row=0, column=1, sticky="nsew", padx=16, pady=18)
        self.main_area.grid_columnconfigure(0, weight=1)
        self.main_area.grid_rowconfigure(3, weight=1)

        self._build_sidebar()
        self._build_main_area()
        self._render_table()
        self._update_metrics()
        self._show_details_empty()

    def _load_visual_assets(self):
        if LOGO_TEXT_PATH.exists():
            logo_source = Image.open(LOGO_TEXT_PATH)
            logo_crop = logo_source.crop((110, 500, 1160, 760))
            self.logo_text_image = ctk.CTkImage(
                light_image=logo_crop,
                dark_image=logo_crop,
                size=(220, 54),
            )

    def _build_sidebar(self):
        header = ctk.CTkFrame(self.sidebar, fg_color="transparent")
        header.grid(row=0, column=0, sticky="ew", padx=24, pady=(18, 10))
        header.grid_columnconfigure(0, weight=1)

        if self.logo_text_image:
            ctk.CTkLabel(
                header,
                text="",
                image=self.logo_text_image,
            ).grid(row=0, column=0, sticky="w", pady=(0, 8))
        else:
            ctk.CTkLabel(
                header,
                text="LeadRanker",
                text_color=COLORS["text"],
                font=ctk.CTkFont(size=24, weight="bold"),
            ).grid(row=0, column=0, sticky="w", pady=(0, 8))

        ctk.CTkLabel(
            header,
            text="Prospecção local com OpenStreetMap.",
            text_color=COLORS["muted"],
            wraplength=230,
            justify="left",
        ).grid(row=1, column=0, sticky="w", pady=(5, 0))

        self.sidebar_controls = ctk.CTkFrame(self.sidebar, fg_color="transparent")
        self.sidebar_controls.grid(row=1, column=0, sticky="new", padx=24)
        self.sidebar_controls.grid_columnconfigure(0, weight=1)

        self.cidade_entry = self._input("Cidade", "Ex: Sorocaba")
        self.nicho_entry = self._input("Nicho", "Ex: clínica estética")
        self.quantidade_entry = self._input("Quantidade", "Ex: 12")
        self.quantidade_entry.insert(0, "12")

        self._section_label("Filtros")
        self.site_filter_menu = self._filter_menu("Site", self.atualizar_filtro_site)
        self.instagram_filter_menu = self._filter_menu("Instagram", self.atualizar_filtro_instagram)
        self.telefone_filter_menu = self._filter_menu("Telefone", self.atualizar_filtro_telefone)

        self._section_label("Fluxo")
        self.buscar_button = self._button("Buscar no OpenStreetMap", self.buscar_leads, primary=True)
        self.analisar_button = self._button("Analisar sites", self.analisar_sites)
        self.mensagens_button = self._button("Gerar mensagens", self.gerar_mensagens)
        self.exportar_button = self._button("Exportar Excel", self.exportar)
        self.whatsapp_button = self._button("Abrir WhatsApp", self.abrir_whatsapp)
        self.limpar_button = self._button("Limpar resultados", self.limpar_resultados)

        status_box = ctk.CTkFrame(self.sidebar, fg_color=COLORS["panel_alt"], corner_radius=8)
        status_box.grid(row=2, column=0, sticky="ew", padx=24, pady=(8, 18))
        status_box.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(
            status_box,
            text="STATUS",
            text_color=COLORS["muted"],
            font=ctk.CTkFont(size=11, weight="bold"),
            anchor="w",
        ).grid(row=0, column=0, sticky="ew", padx=14, pady=(10, 2))

        self.status_label = ctk.CTkLabel(
            status_box,
            text="Pronto. Preencha os campos e busque leads.",
            text_color=COLORS["muted_2"],
            wraplength=220,
            justify="left",
            anchor="w",
        )
        self.status_label.grid(row=1, column=0, sticky="ew", padx=14, pady=(0, 10))

    def _build_main_area(self):
        self.main_area.grid_columnconfigure(0, weight=5, minsize=410)
        self.main_area.grid_columnconfigure(1, weight=3, minsize=300)

        header = ctk.CTkFrame(self.main_area, fg_color="transparent")
        header.grid(row=0, column=0, columnspan=2, sticky="ew")
        header.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(
            header,
            text="Dashboard de prospecção",
            text_color=COLORS["text"],
            font=ctk.CTkFont(size=30, weight="bold"),
        ).grid(row=0, column=0, sticky="w")

        ctk.CTkLabel(
            header,
            text="Busca gratuita usando dados públicos do OpenStreetMap.",
            text_color=COLORS["muted"],
            font=ctk.CTkFont(size=13),
        ).grid(row=1, column=0, sticky="w", pady=(4, 0))

        self.filter_menu = ctk.CTkOptionMenu(
            header,
            values=["Todos"] + STATUSES,
            command=self.filtrar_por_status,
            width=170,
            height=38,
            fg_color=COLORS["panel_alt"],
            button_color=COLORS["panel_hover"],
            button_hover_color=COLORS["border"],
            dropdown_fg_color=COLORS["panel_alt"],
            dropdown_hover_color=COLORS["panel_hover"],
        )
        self.filter_menu.grid(row=0, column=1, rowspan=2, sticky="e")

        self.metrics = ctk.CTkFrame(self.main_area, fg_color="transparent")
        self.metrics.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(22, 18))
        self.metrics.grid_columnconfigure((0, 1, 2, 3), weight=1)

        self.total_metric = self._metric("Total", "0", "Leads encontrados", 0)
        self.hot_metric = self._metric("Quentes", "0", "Alta prioridade", 1)
        self.avg_metric = self._metric("Score médio", "0", "Potencial geral", 2)
        self.called_metric = self._metric("Chamados", "0", "Contato iniciado", 3)

        self.table_card = ctk.CTkFrame(self.main_area, fg_color=COLORS["bg"], corner_radius=10)
        self.table_card.grid(row=2, column=0, rowspan=2, sticky="nsew", padx=(0, 16))
        self.table_card.grid_columnconfigure(0, weight=1)
        self.table_card.grid_rowconfigure(2, weight=1)

        ctk.CTkLabel(
            self.table_card,
            text="Leads priorizados",
            text_color=COLORS["text"],
            font=ctk.CTkFont(size=18, weight="bold"),
        ).grid(row=0, column=0, sticky="w", padx=18, pady=(16, 6))

        table_header = ctk.CTkFrame(self.table_card, fg_color=COLORS["panel_alt"], corner_radius=8)
        table_header.grid(row=1, column=0, sticky="ew", padx=12, pady=(4, 8))

        for index, (label, weight) in enumerate([("Empresa", 5), ("Score", 1), ("Classe", 1)]):
            table_header.grid_columnconfigure(index, weight=weight)
            ctk.CTkLabel(
                table_header,
                text=label,
                text_color=COLORS["muted_2"],
                font=ctk.CTkFont(size=12, weight="bold"),
                anchor="w",
            ).grid(row=0, column=index, sticky="ew", padx=10, pady=9)

        self.table = ctk.CTkScrollableFrame(self.table_card, fg_color=COLORS["bg"])
        self.table.grid(row=2, column=0, sticky="nsew", padx=12, pady=(0, 12))
        self.table.grid_columnconfigure(0, weight=1)

        self.details = ctk.CTkFrame(
            self.main_area,
            width=310,
            fg_color=COLORS["bg"],
            corner_radius=10,
        )
        self.details.grid(row=2, column=1, rowspan=2, sticky="nsew")
        self.details.grid_propagate(False)
        self.details.grid_columnconfigure(0, weight=1)
        self.details.grid_rowconfigure(0, weight=1)

        self.details_content = ctk.CTkScrollableFrame(self.details, fg_color=COLORS["bg"])
        self.details_content.grid(row=0, column=0, sticky="nsew", padx=8, pady=8)
        self.details_content.grid_columnconfigure(0, weight=1)

    def _input(self, label, placeholder):
        ctk.CTkLabel(
            self.sidebar_controls,
            text=label,
            text_color=COLORS["muted_2"],
            font=ctk.CTkFont(size=12, weight="bold"),
        ).pack(anchor="w", pady=(0, 5))

        entry = ctk.CTkEntry(
            self.sidebar_controls,
            placeholder_text=placeholder,
            height=32,
            fg_color=COLORS["panel_alt"],
            border_color=COLORS["border"],
            text_color=COLORS["text"],
            placeholder_text_color=COLORS["muted"],
            corner_radius=8,
        )
        entry.pack(fill="x", pady=(0, 9))
        return entry

    def _filter_menu(self, label, command):
        row = ctk.CTkFrame(self.sidebar_controls, fg_color="transparent")
        row.pack(fill="x", pady=(0, 6))
        row.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(
            row,
            text=label,
            text_color=COLORS["muted_2"],
            font=ctk.CTkFont(size=12, weight="bold"),
        ).grid(row=0, column=0, sticky="w", padx=(0, 8))

        menu = ctk.CTkOptionMenu(
            row,
            values=FILTER_OPTIONS,
            command=command,
            width=94,
            height=28,
            fg_color=COLORS["panel_alt"],
            button_color=COLORS["panel_hover"],
            button_hover_color=COLORS["border"],
            dropdown_fg_color=COLORS["panel_alt"],
            dropdown_hover_color=COLORS["panel_hover"],
        )
        menu.set("Ambos")
        menu.grid(row=0, column=1, sticky="e")
        return menu

    def _section_label(self, text):
        ctk.CTkLabel(
            self.sidebar_controls,
            text=text.upper(),
            text_color=COLORS["muted"],
            font=ctk.CTkFont(size=11, weight="bold"),
        ).pack(anchor="w", pady=(2, 6))

    def _button(self, text, command, primary=False):
        button = ctk.CTkButton(
            self.sidebar_controls,
            text=text,
            command=command,
            height=30,
            corner_radius=8,
            fg_color=COLORS["accent"] if primary else COLORS["panel_alt"],
            hover_color="#dedede" if primary else COLORS["panel_hover"],
            text_color=COLORS["accent_text"] if primary else COLORS["text"],
            font=ctk.CTkFont(size=12, weight="bold"),
        )
        button.pack(fill="x", pady=3)
        return button

    def _metric(self, label, value, helper, column):
        card = ctk.CTkFrame(self.metrics, fg_color="transparent", corner_radius=10)
        card.grid(row=0, column=column, sticky="ew", padx=(0 if column == 0 else 10, 0))

        value_label = ctk.CTkLabel(
            card,
            text=value,
            text_color=COLORS["text"],
            font=ctk.CTkFont(size=26, weight="bold"),
            anchor="w",
        )
        value_label.pack(anchor="w", padx=16, pady=(14, 0))

        ctk.CTkLabel(
            card,
            text=label,
            text_color=COLORS["muted_2"],
            font=ctk.CTkFont(size=13, weight="bold"),
        ).pack(anchor="w", padx=16)

        ctk.CTkLabel(
            card,
            text=helper,
            text_color=COLORS["muted"],
            font=ctk.CTkFont(size=12),
        ).pack(anchor="w", padx=16, pady=(0, 14))

        return value_label

    def _visible_leads(self):
        leads = []
        for index, lead in enumerate(self.leads):
            if self.status_filter != "Todos" and lead.status != self.status_filter:
                continue

            if not self._matches_boolean_filter(bool(lead.site), self.site_filter):
                continue

            if not self._matches_boolean_filter(bool(lead.instagram), self.instagram_filter):
                continue

            if not self._matches_boolean_filter(bool(lead.telefone), self.telefone_filter):
                continue

            leads.append((index, lead))

        return leads

    def _matches_boolean_filter(self, value, selected):
        if selected not in ["Ambos", "Sim"]:
            return not value
        if selected == "Ambos":
            return True
        if selected == "Sim":
            return value
        if selected == "Não":
            return not value
        return True

    def _render_table(self):
        for widget in self.table.winfo_children():
            widget.destroy()

        visible_leads = self._visible_leads()
        if not visible_leads:
            empty_text = (
                "Nenhum lead aparece com os filtros atuais."
                if self.leads
                else "Nenhum lead para mostrar. Busque no OpenStreetMap para começar."
            )
            ctk.CTkLabel(
                self.table,
                text=empty_text,
                text_color=COLORS["muted"],
                font=ctk.CTkFont(size=14),
                wraplength=360,
            ).grid(row=0, column=0, pady=80)
            return

        for row_index, (lead_index, lead) in enumerate(visible_leads):
            is_selected = lead_index == self.selected_index
            fg_color = COLORS["panel_hover"] if is_selected else COLORS["panel_alt"]
            row = ctk.CTkFrame(self.table, fg_color=fg_color, corner_radius=8)
            row.grid(row=row_index, column=0, sticky="ew", pady=4)

            for column, weight in enumerate([5, 1, 1]):
                row.grid_columnconfigure(column, weight=weight)

            values = [
                f"{lead.nome}\n{lead.cidade} - {lead.nicho}",
                str(lead.score),
                lead.classificacao,
            ]

            for column, value in enumerate(values):
                color = CLASS_COLORS.get(lead.classificacao, COLORS["text"]) if column == 2 else COLORS["text"]
                ctk.CTkLabel(
                    row,
                    text=value,
                    text_color=color,
                    justify="left",
                    anchor="w",
                    wraplength=245 if column == 0 else 95,
                    font=ctk.CTkFont(size=12, weight="bold" if column in [1, 2] else "normal"),
                ).grid(row=0, column=column, sticky="ew", padx=10, pady=12)

            row.bind("<Button-1>", lambda event, index=lead_index: self.selecionar_lead(index))
            for child in row.winfo_children():
                child.bind("<Button-1>", lambda event, index=lead_index: self.selecionar_lead(index))

    def _update_metrics(self):
        total = len(self.leads)
        hot = sum(1 for lead in self.leads if lead.classificacao == "Quente")
        avg = round(sum(lead.score for lead in self.leads) / total) if total else 0
        called = sum(1 for lead in self.leads if lead.status != "Novo")

        self.total_metric.configure(text=str(total))
        self.hot_metric.configure(text=str(hot))
        self.avg_metric.configure(text=str(avg))
        self.called_metric.configure(text=str(called))

    def _clear_details(self):
        for widget in self.details_content.winfo_children():
            widget.destroy()

    def _show_details_empty(self):
        self._clear_details()
        ctk.CTkLabel(
            self.details_content,
            text="Detalhes do lead",
            text_color=COLORS["text"],
            font=ctk.CTkFont(size=20, weight="bold"),
        ).grid(row=0, column=0, sticky="w", padx=12, pady=(10, 4))

        ctk.CTkLabel(
            self.details_content,
            text="Selecione um lead na tabela para ver dados, mensagem e ações.",
            text_color=COLORS["muted"],
            wraplength=240,
            justify="left",
        ).grid(row=1, column=0, sticky="w", padx=12, pady=(0, 20))

    def _render_details(self):
        if self.selected_index is None or self.selected_index >= len(self.leads):
            self._show_details_empty()
            return

        lead = self.leads[self.selected_index]
        self._clear_details()

        ctk.CTkLabel(
            self.details_content,
            text=lead.nome,
            text_color=COLORS["text"],
            font=ctk.CTkFont(size=21, weight="bold"),
            wraplength=235,
            justify="left",
        ).grid(row=0, column=0, sticky="w", padx=12, pady=(10, 4))

        ctk.CTkLabel(
            self.details_content,
            text=f"{lead.nicho} em {lead.cidade}",
            text_color=COLORS["muted"],
            wraplength=235,
            justify="left",
        ).grid(row=1, column=0, sticky="w", padx=12, pady=(0, 14))

        score_card = ctk.CTkFrame(self.details_content, fg_color=COLORS["panel_alt"], corner_radius=8)
        score_card.grid(row=2, column=0, sticky="ew", padx=12, pady=(0, 14))
        score_card.grid_columnconfigure((0, 1), weight=1)

        self._detail_pair(score_card, "Score", str(lead.score), 0, 0)
        self._detail_pair(score_card, "Classe", lead.classificacao, 0, 1)
        self._detail_pair(score_card, "Contato", str(lead.score_contato), 1, 0)
        self._detail_pair(score_card, "Presença", str(lead.score_presenca_digital), 1, 1)
        self._detail_pair(score_card, "Oportunidade", str(lead.score_oportunidade), 2, 0)
        self._detail_pair(score_card, "Dados", str(lead.score_qualidade_dados), 2, 1)
        self._detail_pair(score_card, "Dados OSM", str(lead.quantidade_avaliacoes), 3, 0)
        self._detail_pair(score_card, "Tem site", "Sim" if lead.site else "Não", 3, 1)

        info = ctk.CTkFrame(self.details_content, fg_color="transparent")
        info.grid(row=3, column=0, sticky="ew", padx=12, pady=(0, 12))
        info.grid_columnconfigure(0, weight=1)

        lines = [
            f"Telefone: {lead.telefone or 'Não informado'}",
            f"Endereço: {lead.endereco}",
            f"Site: {lead.site or 'Não possui'}",
            f"Site básico: {'Sim' if lead.site_basico else 'Não'}",
        ]
        for index, line in enumerate(lines):
            ctk.CTkLabel(
                info,
                text=line,
                text_color=COLORS["muted_2"],
                anchor="w",
                justify="left",
                wraplength=235,
            ).grid(row=index, column=0, sticky="ew", pady=2)

        extras = [
            f"Instagram: {lead.instagram or 'Não possui'}",
            f"Filial/rede: {'Sim' if lead.filial else 'Não'}",
        ]
        if lead.filial and lead.motivo_filial:
            extras.append(f"Motivo filial: {lead.motivo_filial}")

        for extra_index, line in enumerate(extras, start=len(lines)):
            ctk.CTkLabel(
                info,
                text=line,
                text_color=COLORS["muted_2"],
                anchor="w",
                justify="left",
                wraplength=235,
            ).grid(row=extra_index, column=0, sticky="ew", pady=2)

        motivos_frame = ctk.CTkFrame(self.details_content, fg_color=COLORS["panel_alt"], corner_radius=8)
        motivos_frame.grid(row=4, column=0, sticky="ew", padx=12, pady=(0, 12))
        motivos_frame.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(
            motivos_frame,
            text="Análise do score",
            text_color=COLORS["text"],
            font=ctk.CTkFont(size=14, weight="bold"),
            anchor="w",
        ).grid(row=0, column=0, sticky="ew", padx=12, pady=(10, 4))

        motivos_texto = "\n".join(f"- {motivo}" for motivo in lead.motivos_score)
        ctk.CTkLabel(
            motivos_frame,
            text=motivos_texto or "- Sem motivos calculados.",
            text_color=COLORS["muted_2"],
            justify="left",
            anchor="w",
            wraplength=235,
        ).grid(row=1, column=0, sticky="ew", padx=12, pady=(0, 10))

        status_row = ctk.CTkFrame(self.details_content, fg_color="transparent")
        status_row.grid(row=5, column=0, sticky="ew", padx=12, pady=(0, 12))
        status_row.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(status_row, text="Status", text_color=COLORS["muted_2"]).grid(
            row=0, column=0, sticky="w", padx=(0, 10)
        )
        self.status_menu = ctk.CTkOptionMenu(
            status_row,
            values=STATUSES,
            command=self.atualizar_status_lead,
            fg_color=COLORS["panel_alt"],
            button_color=COLORS["panel_hover"],
            button_hover_color=COLORS["border"],
        )
        self.status_menu.set(lead.status)
        self.status_menu.grid(row=0, column=1, sticky="ew")

        ctk.CTkLabel(
            self.details_content,
            text="Mensagem pronta",
            text_color=COLORS["text"],
            font=ctk.CTkFont(size=15, weight="bold"),
        ).grid(row=6, column=0, sticky="w", padx=12, pady=(0, 6))

        self.message_box = ctk.CTkTextbox(
            self.details_content,
            height=138,
            fg_color=COLORS["panel_alt"],
            text_color=COLORS["text"],
            border_color=COLORS["border"],
            border_width=1,
            corner_radius=8,
            wrap="word",
        )
        self.message_box.grid(row=7, column=0, sticky="ew", padx=12, pady=(0, 14))
        self.message_box.insert("1.0", lead.mensagem_gerada)

        actions = ctk.CTkFrame(self.details_content, fg_color="transparent")
        actions.grid(row=8, column=0, sticky="ew", padx=12, pady=(0, 18))
        actions.grid_columnconfigure((0, 1), weight=1)

        ctk.CTkButton(
            actions,
            text="Copiar mensagem",
            command=self.copiar_mensagem,
            height=40,
            fg_color=COLORS["panel_alt"],
            hover_color=COLORS["panel_hover"],
            text_color=COLORS["text"],
            corner_radius=8,
        ).grid(row=0, column=0, sticky="ew", padx=(0, 8))

        ctk.CTkButton(
            actions,
            text="Abrir WhatsApp",
            command=self.abrir_whatsapp,
            height=40,
            fg_color=COLORS["accent"],
            hover_color="#dedede",
            text_color=COLORS["accent_text"],
            corner_radius=8,
            font=ctk.CTkFont(weight="bold"),
        ).grid(row=0, column=1, sticky="ew")

    def _detail_pair(self, parent, label, value, row, column):
        frame = ctk.CTkFrame(parent, fg_color="transparent")
        frame.grid(row=row, column=column, sticky="ew", padx=14, pady=10)

        ctk.CTkLabel(
            frame,
            text=value,
            text_color=COLORS["text"],
            font=ctk.CTkFont(size=21, weight="bold"),
            anchor="w",
        ).pack(anchor="w")

        ctk.CTkLabel(
            frame,
            text=label,
            text_color=COLORS["muted"],
            font=ctk.CTkFont(size=12),
            anchor="w",
        ).pack(anchor="w")

    def _set_status(self, text):
        self.status_label.configure(text=text)

    def buscar_leads(self):
        # A busca roda em segundo plano para a interface não travar.
        if self.is_loading:
            return

        cidade = self.cidade_entry.get().strip() or "Sorocaba"
        nicho = self.nicho_entry.get().strip() or "clínica estética"

        try:
            quantidade = int(self.quantidade_entry.get().strip() or 12)
        except ValueError:
            quantidade = 12

        quantidade = max(1, min(quantidade, 50))
        self.quantidade_entry.delete(0, "end")
        self.quantidade_entry.insert(0, str(quantidade))

        self.is_loading = True
        self.buscar_button.configure(text="Buscando...", state="disabled")
        self._set_status("Buscando leads no OpenStreetMap...")

        thread = threading.Thread(
            target=self._buscar_leads_worker,
            args=(cidade, nicho, quantidade),
            daemon=True,
        )
        thread.start()

    def _buscar_leads_worker(self, cidade, nicho, quantidade):
        try:
            leads = buscar_leads_openstreetmap(cidade, nicho, quantidade)
            self.after(0, lambda: self._finalizar_busca(cidade, nicho, leads, None))
        except Exception as error:
            self.after(0, lambda erro=error: self._finalizar_busca(cidade, nicho, [], erro))

    def _finalizar_busca(self, cidade, nicho, leads, error):
        self.is_loading = False
        self.buscar_button.configure(text="Buscar no OpenStreetMap", state="normal")

        if error:
            self._set_status(f"Erro na busca: {error}")
            return

        if not leads:
            self.leads = []
            self.selected_index = None
            self._render_table()
            self._render_details()
            self._update_metrics()
            self._set_status(f"Nenhum lead encontrado para {nicho} em {cidade}.")
            return

        self.leads = leads
        self.selected_index = None
        self.status_filter = "Todos"
        self.site_filter = "Ambos"
        self.instagram_filter = "Ambos"
        self.telefone_filter = "Ambos"
        self.filter_menu.set("Todos")
        self.site_filter_menu.set("Ambos")
        self.instagram_filter_menu.set("Ambos")
        self.telefone_filter_menu.set("Ambos")
        self._render_table()
        self._render_details()
        self._update_metrics()
        self._set_status(f"{len(self.leads)} leads encontrados no OpenStreetMap.")

    def analisar_sites(self):
        if self.is_loading:
            return

        if not self.leads:
            self._set_status("Busque leads antes de analisar sites.")
            return

        self.is_loading = True
        self.analisar_button.configure(text="Analisando...", state="disabled")
        self._set_status("Analisando sites encontrados...")

        thread = threading.Thread(target=self._analisar_sites_worker, daemon=True)
        thread.start()

    def _analisar_sites_worker(self):
        try:
            leads = analisar_sites(self.leads)
            self.after(0, lambda: self._finalizar_analise(leads, None))
        except Exception as error:
            self.after(0, lambda erro=error: self._finalizar_analise(self.leads, erro))

    def _finalizar_analise(self, leads, error):
        self.is_loading = False
        self.analisar_button.configure(text="Analisar sites", state="normal")

        if error:
            self._set_status(f"Erro ao analisar sites: {error}")
            return

        self.leads = leads
        self.selected_index = None
        self._render_table()
        self._render_details()
        self._update_metrics()
        self._set_status("Análise concluída. Scores foram recalculados.")

    def gerar_mensagens(self):
        if not self.leads:
            self._set_status("Busque leads antes de gerar mensagens.")
            return

        for lead in self.leads:
            lead.mensagem_gerada = gerar_mensagem(lead)

        self._render_details()
        self._set_status("Mensagens geradas para todos os leads.")

    def selecionar_lead(self, index):
        self.selected_index = index
        self._render_table()
        self._render_details()
        self._set_status(f"Lead selecionado: {self.leads[index].nome}")

    def filtrar_por_status(self, status):
        self.status_filter = status
        self.selected_index = None
        self._render_table()
        self._render_details()
        self._set_status(f"Filtro aplicado: {status}")

    def atualizar_filtro_site(self, valor):
        self.site_filter = valor
        self._aplicar_filtros()

    def atualizar_filtro_instagram(self, valor):
        self.instagram_filter = valor
        self._aplicar_filtros()

    def atualizar_filtro_telefone(self, valor):
        self.telefone_filter = valor
        self._aplicar_filtros()

    def _aplicar_filtros(self):
        self.selected_index = None
        self._render_table()
        self._render_details()
        self._set_status("Filtros aplicados.")

    def atualizar_status_lead(self, status):
        if self.selected_index is None:
            return

        self.leads[self.selected_index].status = status
        self._render_table()
        self._update_metrics()
        self._set_status(f"Status atualizado para: {status}")

    def copiar_mensagem(self):
        if self.selected_index is None:
            self._set_status("Selecione um lead antes de copiar.")
            return

        mensagem = self.message_box.get("1.0", "end").strip()
        self.leads[self.selected_index].mensagem_gerada = mensagem
        self.clipboard_clear()
        self.clipboard_append(mensagem)
        self._set_status("Mensagem copiada para a área de transferência.")

    def exportar(self):
        if not self.leads:
            self._set_status("Busque leads antes de exportar.")
            return

        try:
            caminho = exportar_excel(self.leads)
        except PermissionError:
            self._set_status("Não foi possível exportar. Feche o Excel e tente de novo.")
            return
        except Exception as error:
            self._set_status(f"Erro ao exportar Excel: {error}")
            return

        self._set_status(f"Excel exportado em: {caminho}")

    def limpar_resultados(self):
        if self.is_loading:
            return

        self.leads = []
        self.selected_index = None
        self.status_filter = "Todos"
        self.site_filter = "Ambos"
        self.instagram_filter = "Ambos"
        self.telefone_filter = "Ambos"
        self.filter_menu.set("Todos")
        self.site_filter_menu.set("Ambos")
        self.instagram_filter_menu.set("Ambos")
        self.telefone_filter_menu.set("Ambos")
        self._render_table()
        self._render_details()
        self._update_metrics()
        self._set_status("Resultados limpos. Você pode iniciar uma nova busca.")

    def abrir_whatsapp(self):
        # Abre o WhatsApp com a mensagem pronta, mas o envio continua manual.
        if self.selected_index is None:
            self._set_status("Selecione um lead antes de abrir o WhatsApp.")
            return

        lead = self.leads[self.selected_index]
        telefone = "".join(caractere for caractere in lead.telefone if caractere.isdigit())

        if not telefone:
            self._set_status("O lead selecionado não possui telefone.")
            return

        if hasattr(self, "message_box"):
            lead.mensagem_gerada = self.message_box.get("1.0", "end").strip()

        if not lead.mensagem_gerada:
            lead.mensagem_gerada = gerar_mensagem(lead)

        numero_whatsapp = telefone if telefone.startswith("55") else f"55{telefone}"
        webbrowser.open(f"https://wa.me/{numero_whatsapp}?text={quote(lead.mensagem_gerada)}")
        lead.status = "Chamado"
        self._render_table()
        self._render_details()
        self._update_metrics()
        self._set_status("WhatsApp aberto com mensagem pronta para revisão manual.")


def main():
    app = LeadRankerApp()
    app.mainloop()


if __name__ == "__main__":
    main()
